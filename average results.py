from openpyxl import load_workbook
from collections import defaultdict
import os
import numpy as np 
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

def average_acc_in_excel_files(excel_file_list,filename,num_clients,target_client):
    
    column_titles=["after HiFL","after SelF-HiFL","after Retrain"]
    row_titles=[]
    
    output_wb=Workbook()
    avg_sheet=output_wb.active
    avg_sheet.append([""]+column_titles)
    avg_sheet.title="Average"
    
    std_sheet=output_wb.create_sheet("Standard deviation")
    std_sheet.append([""]+column_titles)

    for row in range(2,num_clients+2):
        data_list=[[] for _ in range(3)]
        for _,file_path in enumerate(excel_file_list):   
            input_wb=load_workbook(file_path)
            input_sheet=input_wb.active

            data_list[0].append(input_sheet.cell(row=row,column=2).value)     # after Hifl
            data_list[1].append(input_sheet.cell(row=row,column=3).value)     # after SelF-HiFL
            data_list[2].append(input_sheet.cell(row=row,column=4).value)     # after Retrain
        hifl_avg=round(np.mean(data_list[0]),2)
        self_avg=round(np.mean(data_list[1]),2)
        retrain_avg=round(np.mean(data_list[2]),2)
        
        hifl_std=round(np.std(data_list[0]),2)
        self_std=round(np.std(data_list[1]),2)
        retrain_std=round(np.std(data_list[2]),2)  

        row_title=input_sheet.cell(row=row,column=1).value
        row_data_avg=[hifl_avg,self_avg,retrain_avg]
        avg_sheet.append([row_title]+row_data_avg)

        row_data_std=[hifl_std,self_std,retrain_std]
        std_sheet.append([row_title]+row_data_std)
        
# avg_sheet--------------------------------------------------------------------------------------------
    red_fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')
    rule=FormulaRule(formula=[f'$A2="{target_client}"'], fill=red_fill, stopIfTrue=True)
    last_col=len(column_titles)+1  
    avg_sheet.conditional_formatting.add(f"A2:{get_column_letter(last_col)}{num_clients+1}", rule)

    for column in avg_sheet.columns:   
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        
        adjusted_width=(max_length+2)* 1.2
        avg_sheet.column_dimensions[column_letter].width=adjusted_width

    
# std_sheet--------------------------------------------------------------------------------------------
    red_fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')
    rule=FormulaRule(formula=[f'$A2="{target_client}"'], fill=red_fill, stopIfTrue=True)
    last_col=len(column_titles)+1  
    std_sheet.conditional_formatting.add(f"A2:{get_column_letter(last_col)}{num_clients+1}", rule)
     


    for column in std_sheet.columns:   
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        
        adjusted_width=(max_length+2)* 1.2
        std_sheet.column_dimensions[column_letter].width=adjusted_width

# --------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------

    avg_sheet['G1']="Avg of clients' acc except target client"
    column_letter='G'
    max_length=len(avg_sheet['G1'].value)
    avg_sheet.column_dimensions[column_letter].width=max_length+1 
    avg_sheet['H1']="target client's acc differences"
    column_letter='H'
    max_length=len(avg_sheet['H1'].value)
    avg_sheet.column_dimensions[column_letter].width=max_length+1
    avg_sheet['F2']="SelF-HiFL acc - HiFL acc"
    avg_sheet['F3']="Retrain acc - HiFL acc"
    column_letter='F'
    max_length=len(avg_sheet['F3'].value)
    avg_sheet.column_dimensions[column_letter].width=max_length+1 
    
    std_sheet['G1']="Avg of clients' acc except target client"
    column_letter='G'
    max_length=len(std_sheet['G1'].value)
    std_sheet.column_dimensions[column_letter].width=max_length+1 
    std_sheet['H1']="target client's acc differences"
    column_letter='H'
    max_length=len(std_sheet['H1'].value)
    std_sheet.column_dimensions[column_letter].width=max_length+1
    std_sheet['F2']="SelF-HiFL acc - HiFL acc"
    std_sheet['F3']="Retrain acc - HiFL acc"
    column_letter='F'
    max_length=len(std_sheet['F3'].value)
    std_sheet.column_dimensions[column_letter].width=max_length+1 
    G2,G3,H2,H3=[],[],[],[]

    for _,file_path in enumerate(excel_file_list):   
        input_wb=load_workbook(file_path)
        input_sheet=input_wb.active
            
        G2.append(input_sheet['G2'].value)
        G3.append(input_sheet['G3'].value)
        H2.append(input_sheet['H2'].value)
        H3.append(input_sheet['H3'].value)
        
    avg_sheet['G2']=np.round(np.mean(G2,axis=0),2)
    avg_sheet['G3']=np.round(np.mean(G3,axis=0),2)
    avg_sheet['H2']=np.round(np.mean(H2,axis=0),2)
    avg_sheet['H3']=np.round(np.mean(H3,axis=0),2)

    std_sheet['G2']=np.round(np.std(G2,axis=0),2)
    std_sheet['G3']=np.round(np.std(G3,axis=0),2)
    std_sheet['H2']=np.round(np.std(H2,axis=0),2)
    std_sheet['H3']=np.round(np.std(H3,axis=0),2)

    thin_border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    
    for row in [1,2,3]:        
        for col in [6,7,8]:      
            avg_sheet.cell(row=row, column=col).border=thin_border
            std_sheet.cell(row=row, column=col).border=thin_border

    for row in avg_sheet.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')

    for row in std_sheet.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')
    output_wb.save(filename)
        
num_clients=4
target_client='client_3'
filename=fr"Average-Standard deviation (training acc).xlsx"
excel_file_list=[fr"./1.xlsx",fr"./2.xlsx"]
average_acc_in_excel_files(excel_file_list,filename,num_clients,target_client)

###############################################################################################################

from openpyxl import load_workbook
from collections import defaultdict
import os
import numpy as np 
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side
from openpyxl.utils import get_column_letter
from scipy import stats


def average_server_target_acc(excel_file_list,filename):
    
    column_titles=["after HiFL","after SelF-HiFL","after Retrain"]
    row_titles=[]
    
    output_wb=Workbook()
    avg_sheet=output_wb.active
    avg_sheet.append([""]+column_titles)
    avg_sheet.title="Average"
    
    std_sheet=output_wb.create_sheet("Standard deviation")
    std_sheet.append([""]+column_titles)

    server=[]
    target=[]
    for _,file_path in enumerate(excel_file_list):   
        input_wb=load_workbook(file_path)
        input_sheet=input_wb.active
        server.append([cell.value for row in input_sheet.iter_rows(min_row=2,max_row=2,min_col=2,max_col=4) for cell in row])
        target.append([cell.value for row in input_sheet.iter_rows(min_row=3,max_row=3,min_col=2,max_col=4) for cell in row])   

    avg_server=np.round(np.mean(server,axis=0),2)
    avg_target=np.round(np.mean(target,axis=0),2)
    std_server=np.round(np.std(server,axis=0),2)
    std_target=np.round(np.std(target,axis=0),2)
    
# avg_sheet-------------------------------------------
    row_title=input_sheet.cell(row=2,column=1).value
    row_data_avg=list(avg_server)
    avg_sheet.append([row_title]+row_data_avg)
    row_title=input_sheet.cell(row=3,column=1).value
    row_data_avg=list(avg_target)
    avg_sheet.append([row_title]+row_data_avg)
    
# std_sheet-------------------------------------------  
    row_title=input_sheet.cell(row=2,column=1).value
    row_data_std=list(std_server)
    std_sheet.append([row_title]+row_data_std)
    row_title=input_sheet.cell(row=3,column=1).value
    row_data_std=list(std_target)
    std_sheet.append([row_title]+row_data_std)

# T-test-------------------------------------------
    avg_sheet['E1']="T-test (between SelF and Retrain)"
    alpha=0.05
    _,p_value=stats.ttest_rel(np.array(server)[:,1],np.array(target)[:,1])
    if p_value<alpha:
        avg_sheet['E2']="significant difference"   
    else:
       avg_sheet['E2']="no significant difference"  
        
    _,p_value=stats.ttest_rel(np.array(server)[:,2],np.array(target)[:,2])
    if p_value<alpha:
        avg_sheet['E3']="significant difference"   
    else:
       avg_sheet['E3']="no significant difference" 
    
# adjusted_width----------------------------------------
    for column in avg_sheet.columns:   
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        adjusted_width=(max_length+2)* 1.2
        avg_sheet.column_dimensions[column_letter].width=adjusted_width

    for column in std_sheet.columns:   
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        adjusted_width=(max_length+2)* 1.2
        std_sheet.column_dimensions[column_letter].width=adjusted_width
    
#alignment(center)    
    for row in avg_sheet.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')

    for row in std_sheet.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')
    output_wb.save(filename)

# Output --------------------------------------------------------------------------    
main_folder=os.getcwd()
excel_name="accu on target train, server test.xlsx"  

excel_file_list=[]
for folder in os.listdir(main_folder):
    folder_path=os.path.join(main_folder,folder)
    if os.path.isdir(folder_path):
        excel_path=os.path.join(folder_path, excel_name)
        if os.path.exists(excel_path):
            wb=load_workbook(filename=excel_path,read_only=True)
            excel_file_list.append(excel_path)

filename=fr"Avg-Std (target-server).xlsx"
average_server_target_acc(excel_file_list,filename)

######################################################################################################

from openpyxl import load_workbook
from collections import defaultdict
import os
import numpy as np 
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side
from openpyxl.utils import get_column_letter
from scipy import stats


def avg_std_diff(excel_file_list,filename):
    
    column_titles=["Avg of clients' acc except target client","target client's acc differences"]
    row_titles=[]
    
    output_wb=Workbook()
    avg_sheet=output_wb.active
    avg_sheet.append([""]+column_titles)
    avg_sheet.title="Average"
    
    std_sheet=output_wb.create_sheet("Standard deviation")
    std_sheet.append([""]+column_titles)

    SelF_HiFL=[]                 # difference between of self's acc and hifl's acc 
    Retrain_HiFL=[]                # difference retrain of self's and hifl's acc  
    
 # comupting avg and std-------------------------------------------
    for _,file_path in enumerate(excel_file_list):   
        input_wb=load_workbook(file_path)
        input_sheet=input_wb.active
        SelF_HiFL.append([cell.value for row in input_sheet.iter_rows(min_row=2,max_row=2,min_col=7,max_col=8) for cell in row])
        Retrain_HiFL.append([cell.value for row in input_sheet.iter_rows(min_row=3,max_row=3,min_col=7,max_col=8) for cell in row])   
    avg_SelF_HiFL=np.round(np.mean(SelF_HiFL,axis=0),2)
    avg_Retrain_HiFL=np.round(np.mean(Retrain_HiFL,axis=0),2)
    std_SelF_HiFL=np.round(np.std(SelF_HiFL,axis=0),2)
    std_Retrain_HiFL=np.round(np.std(Retrain_HiFL,axis=0),2)   
    
# avg_sheet-------------------------------------------
    row_title=input_sheet.cell(row=2,column=6).value
    row_data_avg=list(avg_SelF_HiFL)
    avg_sheet.append([row_title]+row_data_avg)
    row_title=input_sheet.cell(row=3,column=6).value
    row_data_avg=list(avg_Retrain_HiFL)
    avg_sheet.append([row_title]+row_data_avg)
    
# std_sheet-------------------------------------------  
    row_title=input_sheet.cell(row=2,column=6).value
    row_data_std=list(std_SelF_HiFL)
    std_sheet.append([row_title]+row_data_std)
    row_title=input_sheet.cell(row=3,column=6).value
    row_data_std=list(std_Retrain_HiFL)
    std_sheet.append([row_title]+row_data_std)
    
# T-test-------------------------------------------
    avg_sheet['A4']="T-test"
    alpha=0.05
    _,p_value=stats.ttest_rel(np.array(SelF_HiFL)[:,0],np.array(Retrain_HiFL)[:,0])
    if p_value<alpha:
        avg_sheet['B4']="significant difference"   
    else:
       avg_sheet['B4']="no significant difference"  
        
    _,p_value=stats.ttest_rel(np.array(SelF_HiFL)[:,1],np.array(Retrain_HiFL)[:,1])
    if p_value<alpha:
        avg_sheet['C4']="significant difference"   
    else:
       avg_sheet['C4']="no significant difference" 
   
# adjusted_width----------------------------------------
    for column in avg_sheet.columns:   
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        adjusted_width=(max_length+2)* 1.2
        avg_sheet.column_dimensions[column_letter].width=adjusted_width

    for column in std_sheet.columns:   
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        adjusted_width=(max_length+2)* 1.2
        std_sheet.column_dimensions[column_letter].width=adjusted_width
    
#alignment(center)    
    for row in avg_sheet.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')

    for row in std_sheet.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')

    output_wb.save(filename)
        
# OutPut------------------------------------------------------------------------------
main_folder=os.getcwd()
excel_name="train accuracy of clients.xlsx"  

excel_file_list=[]
for folder in os.listdir(main_folder):
    folder_path=os.path.join(main_folder,folder)
    if os.path.isdir(folder_path):
        excel_path=os.path.join(folder_path, excel_name)
        if os.path.exists(excel_path):
            wb=load_workbook(filename=excel_path,read_only=True)
            excel_file_list.append(excel_path)
filename=fr"Avg_Std (differences).xlsx"
avg_std_diff(excel_file_list,filename)
