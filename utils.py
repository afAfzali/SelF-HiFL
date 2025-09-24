import numpy as np 
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import msvcrt
import matplotlib.pyplot as plt
from model.initialize_model import create
from tensorflow.keras.utils import to_categorical
from sklearn.metrics import confusion_matrix,ConfusionMatrixDisplay
from collections import Counter



def average_weights(w,sample_num):
    """
    w: list of lists
    sample_num: list of dataset sizes
    return: list
    """
    total_sample_num=sum(sample_num)
    a=[]
    avg_w=[]
    for i in range(len(w)):
        t=[]
        for j in range(len(w[0])):
            t.append(w[i][j]*(sample_num[i]/total_sample_num))
        a.append(t) 
    avg_w=[sum(k) for k in zip(*a)]  
    return avg_w

def sum_list(a,b):
    result=[]
    for x,y in zip(a,b):
        if isinstance(x,list) and isinstance(y,list):
            result.append(sum_lists(x,y))  
        else:
            result.append(x+y) 
    return result

def l2_norm(x):
    flat=flatten(x)
    return sum(x**2 for x in flat)**0.5

def flatten(x):
    flat_list=np.concatenate([a.flatten() for a in x])
    return flat_list

def multiply(ratio,a):
    result=[]
    for x in a:
        if isinstance(x,list):
            result.append(multiply(ratio,a))
        else:
            result.append(ratio*x)
    return result

def compare(a,b):
    for x,y in zip(a,b):
        if not np.array_equal(x,y):
            return False
    return True
        
def check_target_client_existence(edges,client_name):
    for edge in edges:
        if client_name in edge.cnames:
            target_edge=edge.name
            edge.calibrating_cnames=list(set(edge.cnames)-set([client_name]))
        else: 
            edge.calibrating_cnames=edge.cnames
    print(f"{client_name} belongs to {target_edge}")
    return target_edge

def distribution_dataset(data_structure,num_labels,flag,file_name):
    if isinstance(data_structure,list):
        if flag=="train":
            print("\n\t\t\t\ttrain:",file=file_name)
            for client in data_structure:
                n_labels=[0]*num_labels
                for i,j in client.x:
                    n_labels[np.argmax(j.numpy())]+=1
                client.train_label_frequency=n_labels
                print("--------------------",file=file_name)
                print( client.name,":",n_labels,"-->",np.sum(n_labels),file=file_name)      
            print('\n-----------------------------------------------------------------',file=file_name)
        else:
            print("\t\t\t\ttest:",file=file_name)
            for client in data_structure:
                n_labels=[0]*num_labels
                for i,j in client.y:
                    n_labels[np.argmax(j.numpy())]+=1
                print("--------------------",file=file_name)
                print( client.name,":",n_labels,"-->",np.sum(n_labels),file=file_name)
            print('\n=================================================================',file=file_name)
    else: 
        print('\n=================================================================',file=file_name)
        print("\t\t\t\tserver test:",file=file_name)
        n_labels=[0]*num_labels
        for i,j in data_structure.y:
            n_labels[np.argmax(j.numpy())]+=1
        print(" server:",n_labels,"-->",np.sum(n_labels),file=file_name) 
        print('\n=================================================================',file=file_name)
        
def calculate_label_frequencies(edges,clients,file_name):
    labels_frequencies=[]
    for edge in edges:
        for client_name in edge.cnames:
            index=int(client_name.split('_')[1])-1
            labels_frequencies.append(clients[index].train_label_frequency)
    sum_per_label=np.sum(labels_frequencies,axis=0).tolist()
    print(f"\t\t\ttrain labels frequencies  (%)\n",file=file_name)
    for edge in edges: 
        print(f"\t\t\t\t** {edge.name} **:",file=file_name)
        for client_name in edge.cnames:
            index=int(client_name.split('_')[1])-1
            result=[np.round((x/y)*100,1) for x, y in zip(clients[index].train_label_frequency,sum_per_label)]
            print("--------------------",file=file_name)
            print(clients[index].name,":",result,file=file_name)
        if edge==edges[-1] and client_name==edge.cnames[-1]:
            continue  
        print('\n-----------------------------------------------------------------',file=file_name)
        
def plot_image(idx,X,Y,dataset):
    image=X[idx]
    plt.figure(figsize=(5,5))
    if dataset=="mnist":
        plt.imshow(image,cmap='gray_r')   
        plt.title(f'label={np.argmax(Y[idx])}',fontsize=10)    
    elif dataset=="cifar10":
        labels=["airplane","automobile","bird","cat","deer","dog","frog","horse","ship","truck"]
        plt.imshow(image,interpolation='lanczos')   
        plt.title(f'label={labels[np.argmax(Y_train[idx])]}',fontsize=10) 

def create_file(file_name,row_titles,column_titles,data,target_client=None,num_clients=None,delimiter=','):
    wb=Workbook()
    ws=wb.active
    ws.append([""]+column_titles)

    for i,(row_title,row_data) in enumerate(zip(row_titles,data)):
        ws.append([row_title]+row_data)
    
    if target_client:
        red_fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')
        rule=FormulaRule(formula=[f'$A2="{target_client}"'],fill=red_fill,stopIfTrue=True)
        last_col=len(column_titles)+1  
        ws.conditional_formatting.add(f"A2:{get_column_letter(last_col)}{num_clients+1}", rule)
    
    for column in ws.columns:      #auto-adjust column widths
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        
        adjusted_width=(max_length+2)* 1.2
        ws.column_dimensions[column_letter].width=adjusted_width

    for row in ws.iter_rows():
        for cell in row:
                cell.alignment=Alignment(horizontal='center',vertical='center')
    wb.save(file_name)

def show_confusion_matrix(file_name,flag,server,num_labels):                # /  لیبل رو اوکی کنم که برای همه چی جواب بده          
    y_true=[]
    x=[]
    for i,j in server.y:
        x.append(i.numpy())
        y_true.append(np.argmax(j.numpy()))
    y_pred=server.predict(np.array(x))
    y_pred=np.array([np.argmax(y) for y in y_pred])
    cm=confusion_matrix(np.array(y_true),y_pred)
    cm_normalized=cm.astype('float')/cm.sum(axis=1)[:,np.newaxis]

    diagonal_values=np.diag(cm_normalized)    # correct predictions per class

    plt.figure(figsize=(num_labels,num_labels))
    plt.imshow(cm_normalized,interpolation='none',cmap='Oranges')
    
    # Add black grid lines to separate cells
    for i in range(num_labels + 1):
        plt.axhline(i-0.5,color='black',linewidth=1)
        plt.axvline(i-0.5,color='black',linewidth=1)
    # Annotate cells
    for i in range(num_labels):
        for j in range(num_labels):
            plt.text(j,i,f"{cm[i, j]}\n({cm_normalized[i, j]:.1%})",
                     ha="center",va="center",
                     color="black" if cm_normalized[i, j]>0.5 else "black")
    
    plt.title(f"Confusion Matrix ({flag})- on server's test")
    plt.xlabel('Predicted Label')
    plt.ylabel('True Label')
    class_names=list(range(num_labels))                   #  بعدا بهبود داده بشه 
    plt.xticks(np.arange(num_labels),class_names)
    plt.yticks(np.arange(num_labels),class_names)
    plt.tight_layout()
    plt.savefig('confusion_matrix.png', dpi=300)
    plt.close()
   
    wb=load_workbook(file_name)
    ws=wb.create_sheet(f"CM ({flag})")
    img=Image('confusion_matrix.png')
    ws.add_image(img,'A1')
    
    # Remove gridlines 
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.fill=openpyxl.styles.PatternFill(fill_type=None)
    wb.save(file_name)
    return diagonal_values

def confusion_matrix_difference(file_name,learned_values,unlearned_values,retrained_values,clients,num_labels,target_client):
    wb=load_workbook(file_name)
    ws=wb.create_sheet("differences between CM")
    
    
    #tickmark="\u2714"    # unicode for ✔
    #red_tickmark="\033[31m"+tickmark+"\033[0m"    

    highlight_fill=PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid")
    thin_border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    column_titles=["Retrain - HiFL (%)","Self-HiFL - HiFL (%)"]
    row_titles=[]
    data=[[] for _ in range(num_labels)]
    clients_labels=[[] for _ in range(len(clients))]
    for c_idx,client in enumerate(clients):
        column_titles.append(client.name)
        clients_labels[c_idx]=np.unique([np.argmax(j.numpy()) for _,j in client.x])   
        
    for label in range(num_labels):
        row_titles.append(f"label-{label}")
        data[label].append(100*(np.round(retrained_values[label]-learned_values[label],2)))
        data[label].append(100*(np.round(unlearned_values[label]-learned_values[label],2)))
        for c_idx in range(len(clients)):
            if label in clients_labels[c_idx]:
                data[label].append("✔")
            else:
                data[label].append(" ")

    ws.append([""]+column_titles)
    for i,(row_title,row_data) in enumerate(zip(row_titles,data)):
        ws.append([row_title]+row_data)       

    for cell in ws[1]:  # check each cell in header row  :row1
        if cell.value==target_client:
            region_col=cell.column_letter  
            break
            
    # apply fill to all cells in the column
    for row in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=cell.column,max_col=cell.column):
        for cell in row:
            cell.fill=highlight_fill
        
    for column in ws.columns:      #auto-adjust column widths
        max_length=0
        column_letter=get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value))>max_length:
                    max_length=len(str(cell.value))
            except:
                pass
        
        adjusted_width=(max_length+2)* 1.2
        ws.column_dimensions[column_letter].width=adjusted_width

    for row in ws.iter_rows():
        for cell in row:
            cell.border=thin_border
            cell.alignment=Alignment(horizontal='center',vertical='center')

    wb.save(file_name)

def save_accuracy_changes_to_excel(file_name,target_client,num_clients):
    wb=load_workbook(file_name)
    ws=wb.active
    ws['G1']="Avg of clients' acc except target client"
    column_letter='G'
    max_length=len(ws['G1'].value)
    ws.column_dimensions[column_letter].width=max_length+1 
    ws['H1']="target client's acc differences"
    column_letter='H'
    max_length=len(ws['H1'].value)
    ws.column_dimensions[column_letter].width=max_length+1
    ws['F2']="SelF-HiFL acc - HiFL acc"
    ws['F3']="Retrain acc - HiFL acc"
    column_letter='F'
    max_length=len(ws['F3'].value)
    ws.column_dimensions[column_letter].width=max_length+1 
    
    data=[]
    target_index=int(target_client.split('_')[1]) 
    row_list=list(set(list(range(2,num_clients+2)))-set([target_index+1]))  
    
    for row in row_list: 
        cell1=round(ws.cell(row=row,column=3).value-ws.cell(row=row,column=2).value,2)  # SelF-HiFL acc - Hifl acc
        cell2=round(ws.cell(row=row,column=4).value-ws.cell(row=row,column=2).value,2)  # retrain acc - Hifl acc
        data.append([cell1,cell2])
        
    ws['G2'],ws['G3']=np.round(np.mean(data,axis=0),3)
    ws['H2']=round(ws.cell(row=target_index+1,column=3).value-ws.cell(row=row,column=2).value,3)   # SelF-HiFL acc - Hifl acc
    ws['H3']=round(ws.cell(row=target_index+1,column=4).value-ws.cell(row=row,column=2).value,3)   # retrain acc - Hifl acc
    
    for row in ws.iter_rows():
        for cell in row:
                cell.alignment=Alignment(horizontal='center',vertical='center')
    
    # blue_fill=PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    # for row in [1,2,3]:         
    #     for col in [6,7,8]:      
    #         ws.cell(row=row, column=col).fill=blue_fill
    thin_border=Border(left=Side(style='thin'),right=Side(style='thin'),
        top=Side(style='thin'),bottom=Side(style='thin'))
    
    for row in [1,2,3]:        
        for col in [6,7,8]:      
            ws.cell(row=row, column=col).border=thin_border
    wb.save(file_name)

def counting(a):
    counts=Counter(a)
    sorted_counts=sorted(counts.items(),key=lambda x:x[0])
    return sorted_counts

def seconds_to_scientific(seconds):
    #1 integer digit.
    #2 decimal places.
    power=0
    if seconds>=10.0:
        while seconds>=10.0:
            seconds/=10.0
            power+=1
    elif 0 <seconds<1.0:
        while seconds<1.0:
            seconds*=10.0
            power-=1
    coefficient=round(seconds,2)
    coefficient="{:.2f}".format(coefficient)

    return f"{coefficient}×10^{power}"

# ----------------------------------------------------------------------------
#                      functions for MIA
# ----------------------------------------------------------------------------

def load_model(model,global_r,folder,flag):
    model.load_weights(fr".\{folder}\itr_{global_r}_{flag}.h5")  
    
def save_data(folder,data,sh_id,flag):
    np.save(fr'.\{folder}\shadow-{sh_id+1}-{flag}.npy',data)

def load_data(folder,sh_id,flag):
    return np.load(fr'.\{folder}\shadow-{sh_id+1}-{flag}.npy')

def define_train_data_for_attacks(model,global_r,folder,num_shadows,num_labels,attack_model):
    in_attack_X=[[] for _ in range(num_labels)]
    out_attack_X=[[] for _ in range(num_labels)]
    
    for sh_idx in range(num_shadows):
        load_model(model,global_r,folder,f"final-shadow-{sh_idx+1}")
        
        Y_train=load_data(folder,sh_idx,"Ytrain")
        labels=np.array([np.argmax(y) for y in Y_train])   # train 
        X_train=load_data(folder,sh_idx,"Xtrain")
        prob_list=model.predict(X_train)
        for idx,l in enumerate(labels):
            in_attack_X[l].append(prob_list[idx])
            
        Y_test=load_data(folder,sh_idx,"Ytest")
        labels=np.array([np.argmax(y) for y in Y_test])
        X_test=load_data(folder,sh_idx,"Xtest")
        prob_list=model.predict(X_test)
        for idx,l in enumerate(labels):
            out_attack_X[l].append(prob_list[idx])
    
    train_X_attack=[[] for _ in range(num_labels)]
    train_Y_attack=[[] for _ in range(num_labels)]
   
    for j in range(num_labels):
        if attack_model=="attack_mlp":
            Y_in=to_categorical(np.ones((len(in_attack_X[j]),1)),2)     # in : 1
            Y_out=to_categorical(np.zeros((len(out_attack_X[j]),1)),2)   # out :0  
        elif attack_model=="attack_xgboost":
            Y_in=np.ones((len(in_attack_X[j]),1),dtype=int)    # in : 1
            Y_out=np.zeros((len(out_attack_X[j]),1),dtype=int)  
            
        X_attack=np.vstack((np.array(in_attack_X[j]),np.array(out_attack_X[j])))
        Y_attack=np.vstack((Y_in,Y_out))
        idxs=list(range(len(X_attack)))
        np.random.shuffle(idxs)
        train_X_attack[j]=X_attack[idxs]
        train_Y_attack[j]=Y_attack[idxs]

    #-----------------------------------------
    path=folder.removesuffix('/shadows')
    with open(fr"./{path}/shadow-results.txt", "a") as sh_results: 
        print("\n--> train dataset of attacks: --------------------------------------------\n",file=sh_results)
        for ii in range(len(in_attack_X)):
            n_labels=[0]*num_labels
            for k in range(len(in_attack_X[ii])):
                for ll in range(num_labels):
                    if ll==np.argmax(in_attack_X[ii][k]):
                        n_labels[ll]+=1
            print("attack-",ii,"(in) : ",n_labels,"-->", np.sum(n_labels),file=sh_results)
        print("--------------------------------------------------------------------------",file=sh_results)
              
        for ii in range(len(out_attack_X)):
            n_labels=[0]*num_labels
            for k in range(len(out_attack_X[ii])):
                for ll in range(num_labels):
                    if ll==np.argmax(out_attack_X[ii][k]):
                        n_labels[ll]+=1
            print("attack-",ii,"(out) : ",n_labels,"-->", np.sum(n_labels),file=sh_results)

    with open(fr"./{path}/attack-results.txt", "a") as a_results: 
        print("\n--> distribution of attack datasets (train): --------------------------------------------",file=a_results)
        print("\t    [out  ,  in]  ",file=a_results)
        for i in range(len(train_X_attack)):
            n_labels=[0]*2
            for k in range(len(train_Y_attack[i])):
                for l in range(2):
                    if l==(np.argmax(train_Y_attack[i][k]) if attack_model=="attack_mlp" else train_Y_attack[i][k]):
                        n_labels[l]+=1
            print('attack',i,":",n_labels,"-->",np.sum(n_labels),file=a_results)
    #-----------------------------------------
        
    return train_X_attack,train_Y_attack

def create_excel_attacks_test(num_clients,num_labels,sheet_name,metrics,data,target_client,file_name):
    if sheet_name=="confusion-matrix":
        wb=Workbook()
        del wb["Sheet"]
    else:
        wb=load_workbook(file_name)
    ws=wb.create_sheet(sheet_name)

    # borders
    border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    highlight_fill=PatternFill(start_color="D3D3D3",end_color="D3D3D3",fill_type="solid")

    sections=[{'title':'HiFL',
               'clients':[f'client_{i+1}' for i in range(num_clients)],
               'metrics':metrics,
               'attacks':[f'attack {i}' for i in range(num_labels)]},
              {'title':'Retrained-HiFL',
              'clients':[f'client_{i+1}' for i in range(num_clients)],
             'metrics':metrics,
              'attacks':[f'attack {i}' for i in range(num_labels)]},
              {'title':'SelF-HiFL',
              'clients':[f'client_{i+1}' for i in range(num_clients)],
             'metrics':metrics,
              'attacks':[f'attack {i}' for i in range(num_labels)]}]

    current_row=1  #start
    for section_idx,section in enumerate(sections):
        # section titles
        ws.cell(row=current_row,column=2,value=section['title'])
        end_col=len(section['clients'])*len(section['metrics'])+1
        ws.merge_cells(start_row=current_row,start_column=2,end_row=current_row,end_column=end_col)
        
        for col in range(2,end_col+1):
            ws.cell(row=current_row,column=col).border=border
        
        current_row+=1
        # client headers
        client_col=2
        for client in section['clients']:
            client_cell=ws.cell(row=current_row,column=client_col,value=client)
            end_client_col=client_col+len(section['metrics'])-1
            ws.merge_cells(start_row=current_row,start_column=client_col,end_row=current_row,end_column=end_client_col)
            
            # apply highlight if this is the target client
            if client==target_client:
                for col in range(client_col,end_client_col+1):
                    cell=ws.cell(row=current_row,column=col)
                    cell.border=border
                    cell.fill=highlight_fill
                    
            else:
                for col in range(client_col,end_client_col+1):
                    ws.cell(row=current_row,column=col).border=border
            
            client_col=end_client_col+1
        current_row+=1
        # metric headers
        metric_headers=section['metrics']*len(section['clients'])
        for col_num,metric in enumerate(metric_headers,start=2):
            metric_cell=ws.cell(row=current_row,column=col_num,value=metric)
            metric_cell.border=border
            
            client_index=section['clients'].index(target_client) if target_client in section['clients'] else -1
            if client_index>=0 and (col_num-2)//len(section['metrics'])==client_index:
                metric_cell.fill=highlight_fill
                
        current_row+=1
        # attack rows
        for attack_idx, attack in enumerate(section['attacks']):
            attack_cell=ws.cell(row=current_row,column=1,value=attack)
            attack_cell.border=border
            
            # insert values for each metric column
            for col in range(2,2+len(metric_headers)):
                empty_cell=ws.cell(row=current_row,column=col)
                empty_cell.border=border
                
                # which client this column belongs to
                client_idx=(col-2)//len(section['metrics'])
                metric_idx=(col-2)%len(section['metrics'])

                empty_cell.value=data[section_idx][client_idx][attack_idx][metric_idx]

                if section['clients'][client_idx] == target_client:
                    empty_cell.fill=highlight_fill
            
            current_row+=1
        # add empty row between sections if not last section
        if section!=sections[-1]:
            current_row+=1

    # Adjust column widths
    for col_idx in range(1,ws.max_column+1):
        max_length=0
        column_letter=get_column_letter(col_idx)
        
        # Skip merged columns for width calculation
        is_merged=False
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col==col_idx:
                is_merged=True
                break
        
        if not is_merged:
            for row_idx in range(1,ws.max_row+1):
                cell=ws.cell(row=row_idx,column=col_idx)
                try:
                    if cell.value and len(str(cell.value))>max_length:
                        max_length=len(str(cell.value))
                except:
                    pass
        
        adjusted_width=(max_length+2)*1.2
        ws.column_dimensions[column_letter].width=adjusted_width if max_length>0 else 8.43
        
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')
            
    # Save the file
    wb.save(file_name)


def compare_attack_results(data,file_name):   
    for key in data:
        for sub_list in data[key]:
            for i in range(len(sub_list)):
                if sub_list[i]==0:
                    sub_list[i]="out"
                else:
                    sub_list[i]="in"
                        
    wb=load_workbook(file_name)
    ws=wb.create_sheet("compare attack on models")
    
    border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    
    current_row=1
    for key,value in data.items():
        ws.cell(row=current_row+1,column=1,value=key).border=border
        end_row=len(value)+current_row
        current_row+=1
        ws.merge_cells(start_row=current_row,start_column=1,end_row=end_row,end_column=1)
        
        for row in range(2,end_row+1):
            ws.cell(row=row,column=1).border=border
            
        current_row=end_row
    
    headers=["samples","true label","predict label (HiFL)","predict label (Retrain)","predict label (Self)"]
    
    for col_num,header in enumerate(headers,start=2):
        ws.cell(row=1,column=col_num,value=header)
    
    for col in range(2,len(headers)+2):
        ws.cell(row=1,column=col).border=border
    
    current_row=2
    for key,value in data.items():
        for sample_i,v in enumerate(value):
            empty_cell=ws.cell(row=current_row,column=2)
            empty_cell.border=border
            empty_cell.value=f"s-{sample_i}"
            for col_i,label in enumerate(v):
                empty_cell=ws.cell(row=current_row,column=col_i+3)
                empty_cell.border=border
                empty_cell.value=label 
            current_row+=1

    # Adjust column widths
    for col_idx in range(1,ws.max_column+1):
        max_length=0
        column_letter=get_column_letter(col_idx)
        
        # Skip merged columns for width calculation
        is_merged=False
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col==col_idx:
                is_merged=True
                break
        
        if not is_merged:
            for row_idx in range(1,ws.max_row+1):
                cell=ws.cell(row=row_idx,column=col_idx)
                try:
                    if cell.value and len(str(cell.value))>max_length:
                        max_length=len(str(cell.value))
                except:
                    pass
        
        adjusted_width=(max_length+2)*1.2
        ws.column_dimensions[column_letter].width=adjusted_width if max_length>0 else 8.43
        
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal='center',vertical='center')
    wb.save(file_name)


def test_attacks_on_clients(model,attacks,clients,global_round,folder,flag,num_labels,path,target_client,attack_model):
    
    ''' flag: leraned or unleraned(Self-HiFL) '''
    
    num_clients=len(clients)
    metrics_results=[[] for _ in range(num_clients)]
    confusion_results=[[] for _ in range(num_clients)]

    load_model(model,global_round,folder,flag)
    data={}   # for target client

    if flag=="learned":
        with open(fr"./{path}/attack-results.txt", "a") as a_results: 
            print("\n--> distribution of attack datasets (test) : --------------------------------------------",file=a_results)
    for c_idx,client in enumerate(clients):
        in_test_labels=np.array([np.argmax(j.numpy()) for _,j in client.x])      # client.x is train data of client 
        in_labels_frequency=counting(in_test_labels)    # new       is sorted 
        # print(client.name,":","in_labels_frequency:",in_labels_frequency)
        in_test_labels=in_test_labels.reshape(len(in_test_labels),1)    #   
        in_test_probs=model.predict(np.array([i.numpy() for i,_ in client.x]))
        
        out_test_labels=np.array([np.argmax(j.numpy()) for _,j in client.y])    # client.y is test data of client
        out_labels_frequency=counting(out_test_labels)    # new     is sorted 
        # print(client.name,":","out_labels_frequency:",out_labels_frequency)
        out_test_labels=out_test_labels.reshape(len(out_test_labels),1)    #  
        out_test_probs=model.predict(np.array([i.numpy() for i,_ in client.y]))

        #----------------------------------------------------------- new
        #client_labels=[x[0] for x in in_labels_frequency]  
        in_indices=[]
        out_indices=[]
        for item1,item2 in zip(in_labels_frequency,out_labels_frequency):   # item1[0]: label    item1[1]: frequency 
            if item1[1]<item2[1]:                              
                indices=[index for index,value in enumerate(out_test_labels) if value==item2[0]]
                indices=indices[:item1[1]]
                out_indices.extend(indices)
                indices=[index for index,value in enumerate(in_test_labels) if value==item2[0]]
                in_indices.extend(indices)
                #print("len:",len(indices),"out_indices:",indices)
            elif item1[1]>item2[1]:
                indices=[index for index,value in enumerate(in_test_labels) if value==item1[0]]
                indices=indices[:item2[1]]
                in_indices.extend(indices)
                indices=[index for index,value in enumerate(out_test_labels) if value==item2[0]]
                out_indices.extend(indices)
            elif item1[1]==item2[1]:
                indices=[index for index,value in enumerate(in_test_labels) if value==item1[0]]
                in_indices.extend(indices)
                indices=[index for index,value in enumerate(out_test_labels) if value==item2[0]]
                out_indices.extend(indices)
        in_test_labels=in_test_labels[in_indices]
        in_test_probs=in_test_probs[in_indices]
        out_test_labels=out_test_labels[out_indices]
        out_test_probs=out_test_probs[out_indices]
        #--------------------------------------------------------------        
        
        # label_in=to_categorical(np.ones((client.train_num,1)),2)
        # label_out=to_categorical(np.zeros((client.test_num,1)),2)
        # print("len:",len(indices),"in_indices:",indices)
        if attack_model=="attack_mlp":
            label_in=to_categorical(np.ones((len(in_test_labels),1)),2)     # new    len(in_test_labels)=len(out_test_labels)
            label_out=to_categorical(np.zeros((len(in_test_labels),1)),2)   # new
        elif attack_model=="attack_xgboost":
            label_in=np.ones((len(in_test_labels),1),dtype=int)   
            label_out=np.zeros((len(in_test_labels),1),dtype=int)

        test_labels=np.vstack((in_test_labels,out_test_labels))
        test_probs=np.vstack((in_test_probs,out_test_probs))
        in_out_labels=np.vstack((label_in,label_out))

        test_X_attack=[[] for _ in range(num_labels)]
        test_Y_attack=[[] for _ in range(num_labels)]
            
        for label,prob,category in zip(test_labels,test_probs,in_out_labels):
            test_X_attack[label[0]].append(prob)
            test_Y_attack[label[0]].append(category)
        
        for at_i,attack in enumerate(attacks):
            if len(test_X_attack[at_i])!=0:    
                if attack_model=="attack_mlp":
                    predict_y=attacks[at_i].model.predict(np.array(test_X_attack[at_i]))
                    predict_y=np.array([np.argmax(y) for y in predict_y])
                    true_y=np.array([np.argmax(y) for y in test_Y_attack[at_i]])
                elif attack_model=="attack_xgboost":
                    predict_y=attacks[at_i].grid_search.predict(np.array(test_X_attack[at_i]))
                    true_y=np.array(test_Y_attack[at_i])
                #----- new added
                if client.name==target_client:
                    data[attack.name]=[true_y,predict_y]
                #-----
                cm=confusion_matrix(true_y,predict_y) 
                tn,fp,fn,tp=cm.ravel()  
                confusion_results[c_idx].append([tn,fp,fn,tp])
                accuracy=np.round((tp+tn)/(tp+tn+fp+fn),2)
                if tp+fp>0:
                    precision=np.round(tp/(tp+fp),2)
                else:
                    precision="nan"   # np.nan
                if tp+fn>0:   
                    recall=np.round(tp/(tp+fn),2)
                else:
                    recall="nan"    # np.nan
                if precision=="nan" or recall=="nan" or (precision+recall==0):   #np.isnan(precision) or np.isnan(recall) or
                    f1_score="nan"      #np.nan
                else:
                    f1_score=np.round(2*(precision*recall)/(precision+recall),2)
                metrics_results[c_idx].append([accuracy,precision,recall,f1_score])
            else:
                confusion_results[c_idx].append(["-","-","-","-"])
                metrics_results[c_idx].append(["-","-","-","-"])     

        if flag=="learned":
            with open(fr"./{path}/attack-results.txt", "a") as a_results: 
                print(f"\n--> {client.name}",file=a_results)
                print("\t    [out  ,  in]  ",file=a_results)
                for i in range(len(test_X_attack)):
                    n_labels=[0]*2
                    for k in range(len(test_Y_attack[i])):
                        for l in range(2):
                            if l==np.argmax(test_Y_attack[i][k]):
                                n_labels[l]+=1
                    print('attack',i,":",n_labels,"-->",np.sum(n_labels),file=a_results)
                
    return confusion_results,metrics_results,data

